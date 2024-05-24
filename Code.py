import requests
from openpyxl import Workbook, load_workbook
import os
from googletrans import Translator
from textblob import TextBlob

def search_most_viewed_videos(Anahtar_Kelime, max_video=5):

    api_key = ''
    url = f'https://www.googleapis.com/youtube/v3/search?key={api_key}&part=snippet&q={Anahtar_Kelime}&type=video&order=viewCount&maxResults={max_video}'
    
    response = requests.get(url)
    data = response.json()
    
    video_id = []
    for item in data['items']:
        video_id.append(item['id']['videoId'])
    
    return video_id



Anahtar_Kelime = '' #video ismi giriniz
max_video = 5   #Kaç adet video almak istediğin
video_id = search_most_viewed_videos(Anahtar_Kelime, max_video)

print(video_id)
for video_id in video_id:

    dosya_adi = "yorumlar.xlsx"


    if os.path.exists(dosya_adi):
        Depo = load_workbook(dosya_adi)
        Sayfa = Depo.active

    else:
        Depo = Workbook()
        Sayfa = Depo.active
        Sayfa[f"A1"] = "Video ID"
        for say in range(2, 12):
            Sayfa[f"A{say}"] = f"Yorum {say-1}"
        for say in range(12, 22):
            Sayfa[f"A{say}"] = f"Analiz {say-11}"





    def fetch_youtube_comments(video_id, api_key):
        base_url = "https://www.googleapis.com/youtube/v3/commentThreads"

        params = {
            'part': 'snippet',
            'videoId': video_id,
            'key': api_key,
            'maxResults': 10,  # İstediğiniz kadar yorum almak için isteğe bağlı olarak
        }

        response = requests.get(base_url, params=params)
        data = response.json()

        if 'items' in data:
            comments = []
            for item in data['items']:
                comment = item['snippet']['topLevelComment']['snippet']['textDisplay']
                comments.append(comment)
            return comments
        else:
            return None
        

    api_key = ''

    comments = fetch_youtube_comments(video_id, api_key)

    if comments:
        
        column = Sayfa.max_column + 1

        Sayfa.cell(row=1, column=column, value=video_id)
        for idx, comment in enumerate(comments, 1):
            Sayfa.cell(row=idx+1, column=column, value=comment)

            Depo.save(dosya_adi)

    else:
        print("Yorum bulunamadı.")

    
    def çevirisi(text):
        try:
            translator = Translator()
            ceviri = translator.translate(text, dest='en')
            return ceviri.text
    
        except Exception as hata:
            print("Çeviri yapılırken bir hata oluştu:", hata)
            return text


    Depo = load_workbook(dosya_adi)
    Sayfa = Depo.active

    Sütun = Sayfa.max_column

    for Satir in range(2, 12):

        hücre_içi = Sayfa.cell(row=Satir, column=Sütun).value
            

        çevrilmiş = çevirisi(hücre_içi)
        analiz = TextBlob(çevrilmiş)

        duygu = analiz.sentiment
        print(f"Yorum: {hücre_içi}")
        print(f"Duygu Analizi: {duygu.polarity}")
        if duygu.polarity > 0:
            Sayfa.cell(row=Satir+10, column=Sütun, value=f"Pozitif: {duygu.polarity}")
            print("Pozitif")
        elif duygu.polarity == 0:
            Sayfa.cell(row=Satir+10, column=Sütun, value=f"Nötür: {duygu.polarity}")
            print("Nötr")
        else:
            Sayfa.cell(row=Satir+10, column=Sütun, value=f"Negatif: {duygu.polarity}")
            print("Negatif")
        print()



    Depo.save(dosya_adi)

